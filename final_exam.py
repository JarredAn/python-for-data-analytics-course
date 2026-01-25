print("Jarand5122")

import csv
import openpyxl
from openpyxl.chart import Reference, PieChart
import matplotlib.pyplot as plt
from datetime import datetime

# Function to ask user for five numbers and display total
def askUser():
    # Loop that iterates 5 times to get 5 numbers from user
    # Accumulates the total by adding each number to the running sum
    total = 0
    for i in range(1, 6):
        number = float(input(f"Enter number {i}: "))
        total += number
    
    print(f"\nThe sum of the five numbers is: {total}")
    print()

# Function to ask for income data and append to CSV file
def askIncome():
    file_path = "C:\\finalExam\\final.csv"
    
    # Loop that iterates 5 times to collect name and income from user
    # Each iteration appends one person's data to the CSV file
    for i in range(1, 6):
        name = input(f"Enter name of person {i}: ")
        income = input(f"Enter annual income for {name}: ")
        
        # Append data to CSV file with newline character
        with open(file_path, 'a', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow([name, income])
    
    print("\nData successfully added to final.csv")
    print()

# Function to create Excel pie chart
def excelPie():
    csv_path = "C:\\finalExam\\final.csv"
    excel_path = "C:\\finalExam\\final.xlsx"
    
    # Read CSV data
    names = []
    incomes = []
    
    with open(csv_path, 'r') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            names.append(row[0])
            incomes.append(int(row[1]))  # Cast to int
    
    # Create new Excel workbook
    wb = openpyxl.Workbook()
    # Get the active worksheet
    ws = wb.active
    
    # Write header row
    ws['A1'] = 'Name'
    ws['B1'] = 'Income'
    
    # Write data to worksheet
    for i, (name, income) in enumerate(zip(names, incomes), start=2):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = income
    
    rows = len(names) + 1  # +1 for header
    
    # Creates a reference to the category labels (names) for the pie chart
    myDataLabels = Reference(ws, min_col=1, min_row=2, max_row=rows)
    
    # Creates a reference to the numerical data (incomes) to be plotted
    myData = Reference(ws, min_col=2, min_row=2, max_row=rows)
    
    # Creates a PieChart object to hold the chart
    myPieChart = PieChart()
    
    # Sets the title of the pie chart with student ID and current date
    today = datetime.now().strftime("%B %d, %Y")
    myPieChart.title = f"Jarand5122 {today}"
    
    # Adds the income data to the pie chart
    myPieChart.add_data(myData)
    
    # Sets the category labels for each pie slice
    myPieChart.set_categories(myDataLabels)
    
    # Adds the pie chart to the worksheet at cell D2
    ws.add_chart(myPieChart, "D2")
    
    # Sets the height of the chart to make it readable
    myPieChart.height = 15
    
    # Sets the width of the chart to make it readable
    myPieChart.width = 15
    
    # Saves the workbook as an Excel file
    wb.save(excel_path)
    
    print(f"Excel pie chart created and saved to {excel_path}")
    print()

# Function to create matplotlib vertical bar graph
def verticalBar():
    csv_path = "C:\\finalExam\\final.csv"
    
    # Read CSV data
    names = []
    incomes = []
    
    with open(csv_path, 'r') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            names.append(row[0])
            incomes.append(int(row[1]))
    
    # Create vertical bar chart
    plt.bar(names, incomes)
    plt.xlabel('Names')
    plt.ylabel('Annual Income')
    
    # Set title with student ID and current date
    today = datetime.now().strftime("%B %d, %Y")
    plt.title(f"Jarand5122 {today}")
    
    # Rotate x-axis labels for readability
    plt.xticks(rotation=45, ha='right')
    
    # Adjust layout to prevent label cutoff
    plt.tight_layout()
    
    # Display the chart
    plt.show()

# Main program execution
def main():
    print("Final Exam Program")
    print("=" * 50)
    print()
    
    # Call askUser function
    print("Part 1: Sum of Five Numbers")
    print("-" * 50)
    askUser()
    
    # Call askIncome function
    print("Part 2: Enter Income Data")
    print("-" * 50)
    askIncome()
    
    # Call excelPie function
    print("Part 3: Creating Excel Pie Chart")
    print("-" * 50)
    excelPie()
    
    # Call verticalBar function
    print("Part 4: Creating Vertical Bar Graph")
    print("-" * 50)
    verticalBar()
    
    print("\nAll tasks completed successfully!")

# Run the main function
if __name__ == "__main__":
    main()
